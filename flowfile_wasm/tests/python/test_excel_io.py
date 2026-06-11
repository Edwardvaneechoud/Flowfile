"""Excel read path (execute_read_excel / list_excel_sheets).

Workbooks are built in-test via openpyxl into BytesIO — the same pure-Python
engine the browser micropip-installs (requirements.txt pins must match the
ensurePyPackages pins in flow-store.ts). The memoryview cases mirror the
Pyodide bridge, which passes `_temp_bytes.to_py()` (a memoryview), not bytes.
"""
import io

import openpyxl
import pytest

import engine


def make_workbook(sheets: dict[str, list[list]]) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


@pytest.fixture
def people_xlsx() -> bytes:
    return make_workbook(
        {
            "people": [["name", "age"], ["alice", 30], ["bob", 25]],
            "empty_ish": [["x"], [1]],
        }
    )


def read_settings(**table_settings) -> dict:
    ts = {"file_type": "excel", "has_headers": True, **table_settings}
    return {"received_file": {"file_type": "excel", "table_settings": ts}}


def test_read_excel_first_sheet(people_xlsx):
    r = engine.execute_read_excel(1, people_xlsx, read_settings())
    assert r["success"] is True
    assert [c["name"] for c in r["schema"]] == ["name", "age"]
    df = engine.get_lazyframe(1).collect()
    assert df["name"].to_list() == ["alice", "bob"]
    assert df["age"].to_list() == [30, 25]


def test_read_excel_explicit_sheet(people_xlsx):
    r = engine.execute_read_excel(1, people_xlsx, read_settings(sheet_name="empty_ish"))
    assert r["success"] is True
    assert [c["name"] for c in r["schema"]] == ["x"]


def test_read_excel_without_headers(people_xlsx):
    r = engine.execute_read_excel(1, people_xlsx, read_settings(has_headers=False))
    assert r["success"] is True
    df = engine.get_lazyframe(1).collect()
    assert len(df) == 3  # header row becomes data
    assert df.row(0)[0] == "name"


def test_read_excel_start_row_promotes_header():
    content = make_workbook({"s": [["junk", None], ["name", "age"], ["alice", 30]]})
    r = engine.execute_read_excel(1, content, read_settings(start_row=1))
    assert r["success"] is True
    df = engine.get_lazyframe(1).collect()
    assert df.columns == ["name", "age"]
    assert len(df) == 1


def test_read_excel_accepts_memoryview(people_xlsx):
    r = engine.execute_read_excel(1, memoryview(people_xlsx), read_settings())
    assert r["success"] is True


def test_read_excel_corrupt_bytes_reports_error():
    r = engine.execute_read_excel(1, b"not a zip archive", read_settings())
    assert r["success"] is False
    assert "node #1" in r["error"]


def test_list_excel_sheets(people_xlsx):
    r = engine.list_excel_sheets(people_xlsx)
    assert r == {"success": True, "sheets": ["people", "empty_ish"]}


def test_list_excel_sheets_memoryview(people_xlsx):
    assert engine.list_excel_sheets(memoryview(people_xlsx))["success"] is True


def test_list_excel_sheets_corrupt():
    r = engine.list_excel_sheets(b"garbage")
    assert r["success"] is False
    assert "workbook" in r["error"].lower()


def output_excel_settings(name="result.xlsx", sheet="Data") -> dict:
    return {
        "output_settings": {
            "name": name,
            "file_type": "excel",
            "table_settings": {"file_type": "excel", "sheet_name": sheet},
        }
    }


def test_output_excel_round_trip(people_xlsx):
    engine.execute_read_excel(1, people_xlsx, read_settings())
    r = engine.execute_output(2, 1, output_excel_settings())

    assert r["success"] is True
    dl = r["download"]
    assert dl["content_kind"] == "binary"
    assert dl["content"] == ""
    assert dl["row_count"] == 2
    assert dl["mime_type"].endswith("sheet")

    raw = engine.take_output_binary(2)
    assert raw[:2] == b"PK"  # xlsx is a zip archive
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    assert wb.sheetnames == ["Data"]
    rows = list(wb["Data"].values)
    assert rows[0] == ("name", "age")
    assert rows[1] == ("alice", 30)

    # One-shot pull: the registry entry is gone
    assert engine.take_output_binary(2) is None


def test_output_excel_rerun_replaces_stale_bytes(people_xlsx):
    engine.execute_read_excel(1, people_xlsx, read_settings())
    engine.execute_output(2, 1, output_excel_settings())
    # JS pull never happened; a re-run as CSV must still drop the staged bytes
    r = engine.execute_output(2, 1, {"output_settings": {"name": "o.csv", "file_type": "csv"}})
    assert r["success"] is True
    assert r["download"]["content_kind"] == "text"
    assert engine.take_output_binary(2) is None

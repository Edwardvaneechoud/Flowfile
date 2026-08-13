import datetime
import multiprocessing
import os
import time

import fastexcel
import openpyxl
import polars as pl
import pytest
from openpyxl.utils.exceptions import IllegalCharacterError

from shared.excel_writer import write_excel_output


def _parallel_update(path, sheet, barrier, delay):
    barrier.wait()
    time.sleep(delay)
    write_excel_output(pl.DataFrame({sheet: [1, 2]}), path, sheet_name=sheet, write_mode="update")


def _seed(path, sheets):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in sheets:
        wb.create_sheet(name).append(["seeded", name])
    wb.save(path)


def test_update_creates_file_when_missing(tmp_path):
    """update on a missing path writes a fresh single-sheet workbook."""
    path = tmp_path / "out.xlsx"
    write_excel_output(pl.DataFrame({"a": [1, 2]}), str(path), sheet_name="Data", write_mode="update")
    assert fastexcel.read_excel(str(path)).sheet_names == ["Data"]
    assert pl.read_excel(str(path), sheet_name="Data").to_dicts() == [{"a": 1}, {"a": 2}]


def test_update_preserves_other_sheets(tmp_path):
    """Writing sheet B in update mode leaves the untouched sheet A in the workbook."""
    path = tmp_path / "book.xlsx"
    write_excel_output(pl.DataFrame({"keep": ["a"]}), str(path), sheet_name="A")
    write_excel_output(pl.DataFrame({"new": [1]}), str(path), sheet_name="B", write_mode="update")
    assert fastexcel.read_excel(str(path)).sheet_names == ["A", "B"]
    assert pl.read_excel(str(path), sheet_name="A").to_dicts() == [{"keep": "a"}]
    assert pl.read_excel(str(path), sheet_name="B").to_dicts() == [{"new": 1}]


def test_update_replaces_target_sheet_at_same_position(tmp_path):
    """Replacing a middle sheet keeps its tab index and does not duplicate it."""
    path = tmp_path / "ordered.xlsx"
    _seed(path, ["A", "Target", "C"])
    write_excel_output(pl.DataFrame({"v": [7]}), str(path), sheet_name="Target", write_mode="update")
    assert fastexcel.read_excel(str(path)).sheet_names == ["A", "Target", "C"]
    assert pl.read_excel(str(path), sheet_name="Target").to_dicts() == [{"v": 7}]


def test_update_preserves_formatting_and_formulas(tmp_path):
    """Fonts and formula text on an untouched sheet survive the update round trip."""
    path = tmp_path / "styled.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Styled"
    ws["A1"] = "header"
    ws["A1"].font = openpyxl.styles.Font(bold=True)
    ws["A2"] = "=A1+1"
    wb.create_sheet("Data")
    wb.save(path)

    write_excel_output(pl.DataFrame({"v": [1]}), str(path), sheet_name="Data", write_mode="update")

    reloaded = openpyxl.load_workbook(path)
    assert reloaded["Styled"]["A1"].font.bold is True
    assert reloaded["Styled"]["A2"].value == "=A1+1"


def test_create_raises_and_leaves_file_untouched(tmp_path):
    """create against an existing file raises FileExistsError without rewriting a byte."""
    path = tmp_path / "existing.xlsx"
    write_excel_output(pl.DataFrame({"a": [1]}), str(path), sheet_name="A")
    before = path.read_bytes()
    with pytest.raises(FileExistsError, match="already exists"):
        write_excel_output(pl.DataFrame({"b": [2]}), str(path), sheet_name="B", write_mode="create")
    assert path.read_bytes() == before


@pytest.mark.parametrize("write_mode", ["new file", "error", "", None])
def test_unknown_write_mode_falls_back_to_overwrite(tmp_path, write_mode):
    """Modes excel never honoured keep replacing the whole workbook, exactly as before."""
    path = tmp_path / "legacy.xlsx"
    write_excel_output(pl.DataFrame({"old": [1]}), str(path), sheet_name="A")
    write_excel_output(pl.DataFrame({"new": [2]}), str(path), sheet_name="B", write_mode=write_mode)
    assert fastexcel.read_excel(str(path)).sheet_names == ["B"]


def test_update_normalizes_nested_and_tz_dtypes(tmp_path):
    """Nested and tz-aware values land exactly as write_excel renders them."""
    path = tmp_path / "dtypes.xlsx"
    df = pl.DataFrame({
        "lst": [[1, 2]],
        "meta": [{"a": 1}],
        "ts": pl.Series([datetime.datetime(2024, 1, 2, 3, 4, 5, tzinfo=datetime.timezone.utc)]).cast(
            pl.Datetime("us", "UTC")
        ),
    })
    write_excel_output(pl.DataFrame({"keep": ["a"]}), str(path), sheet_name="A")
    write_excel_output(df, str(path), sheet_name="B", write_mode="update")
    assert pl.read_excel(str(path), sheet_name="B").to_dicts() == [
        {"lst": "[1, 2]", "meta": "{'a': 1}", "ts": datetime.datetime(2024, 1, 2, 3, 4, 5)}
    ]


def test_update_leaves_no_temp_file_and_preserves_original_on_failure(tmp_path):
    """A failing update keeps the original workbook readable and cleans up its temp file."""
    path = tmp_path / "fragile.xlsx"
    write_excel_output(pl.DataFrame({"keep": ["a"]}), str(path), sheet_name="A")
    before = path.read_bytes()
    with pytest.raises(IllegalCharacterError):
        write_excel_output(pl.DataFrame({"blob": [b"\x00\x01"]}), str(path), sheet_name="B", write_mode="update")
    assert path.read_bytes() == before
    assert pl.read_excel(str(path), sheet_name="A").to_dicts() == [{"keep": "a"}]
    assert list(tmp_path.glob("*.tmp")) == []


@pytest.mark.parametrize("preseeded", [True, False])
def test_parallel_update_writers_keep_every_sheet(tmp_path, preseeded):
    """Concurrent update writers to one workbook serialize instead of losing sheets or corrupting it."""
    path = str(tmp_path / "parallel.xlsx")
    names = [f"S{i}" for i in range(6)]
    expected = set(names)
    if preseeded:
        write_excel_output(pl.DataFrame({"base": [1]}), path, sheet_name="Base")
        expected.add("Base")
    ctx = multiprocessing.get_context("spawn")
    barrier = ctx.Barrier(len(names))
    # Staggered starts on a shared epoch: arrivals spread across several lock hand-offs,
    # which is the interleaving that loses sheets when the locking is broken.
    processes = [
        ctx.Process(target=_parallel_update, args=(path, name, barrier, index * 0.03))
        for index, name in enumerate(names)
    ]
    for process in processes:
        process.start()
    for process in processes:
        process.join(timeout=120)
    assert [process.exitcode for process in processes] == [0] * len(names)
    assert set(fastexcel.read_excel(path).sheet_names) == expected
    if os.name != "nt":
        assert list(tmp_path.glob("*.lock")) == []

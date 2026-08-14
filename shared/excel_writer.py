"""Polars-DataFrame -> Excel writer with sheet-level update support, shared by the worker and core.

``overwrite`` and ``create`` go straight through ``DataFrame.write_excel`` (xlsxwriter), so the
default path stays byte-identical to what Flowfile has always written. ``update`` is the only mode
needing a read-modify-write round trip — something xlsxwriter cannot do at all — so it reopens the
workbook with openpyxl, replaces just the target sheet, and rewrites the file atomically.

Two openpyxl round-trip caveats apply to ``update`` only:
- embedded images anywhere in the workbook are dropped;
- cached formula results are cleared (the formula text survives), so non-Excel readers see empty
  cells on preserved sheets until Excel recomputes and saves the workbook.

Flows run nodes on a thread pool, so two output nodes can target the same workbook concurrently —
with ``update`` that is the intended way to fill several tabs of one file. ``update`` and ``create``
therefore serialize on a per-workbook advisory lock (a ``<path>.lock`` sidecar, removed on release),
and every ``update`` write lands via an atomic ``os.replace``, so concurrent writers merge
sequentially instead of losing sheets or corrupting the file. ``overwrite`` keeps its historical
unlocked last-writer-wins behavior.
"""

from __future__ import annotations

import errno
import os
import shutil
import time
from contextlib import contextmanager
from typing import TYPE_CHECKING, Final

import polars as pl

if TYPE_CHECKING:
    from collections.abc import Iterator

    from openpyxl import Workbook

EXCEL_WRITE_MODES: Final = ("overwrite", "create", "update")

_NESTED_DTYPES: Final = (pl.List, pl.Array, pl.Struct, pl.Object)

_LOCK_TIMEOUT_S: Final = 300.0
_LOCK_POLL_S: Final = 0.05

# Contention looks like EACCES/EAGAIN (msvcrt) or EWOULDBLOCK==EAGAIN (flock); anything else
# (ENOLCK/EOPNOTSUPP on network filesystems, EBADF) means locking is broken and must fail fast.
_CONTENTION_ERRNOS: Final = frozenset({errno.EACCES, errno.EAGAIN, errno.EDEADLK})

if os.name == "nt":
    import msvcrt

    def _try_lock(fd: int) -> bool:
        try:
            msvcrt.locking(fd, msvcrt.LK_NBLCK, 1)
            return True
        except OSError as exc:
            if exc.errno in _CONTENTION_ERRNOS:
                return False
            raise

    def _unlock(fd: int) -> None:
        msvcrt.locking(fd, msvcrt.LK_UNLCK, 1)

else:
    import fcntl

    def _try_lock(fd: int) -> bool:
        try:
            fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
            return True
        except OSError as exc:
            if exc.errno in _CONTENTION_ERRNOS:
                return False
            raise

    def _unlock(fd: int) -> None:
        fcntl.flock(fd, fcntl.LOCK_UN)


def _is_live_lockfile(fd: int, lock_path: str) -> bool:
    try:
        current = os.stat(lock_path)
        held = os.fstat(fd)
    except OSError:
        return False
    return (current.st_dev, current.st_ino) == (held.st_dev, held.st_ino)


@contextmanager
def _workbook_lock(path: str) -> Iterator[None]:
    """Serialize read-modify-write access to *path* across processes.

    Safe-unlink lockfile protocol: after acquiring the flock, verify the fd still is the live
    lockfile — a releaser unlinks *before* unlocking, so a waiter that acquires a stale inode
    detects the mismatch and reopens instead of entering a second lock domain. On Windows an
    open file cannot be unlinked, which makes the plain unlock-close-unlink order safe there.
    The OS releases the lock if the holder dies, so a killed write never wedges later runs.
    """
    lock_path = os.path.realpath(path) + ".lock"
    deadline = time.monotonic() + _LOCK_TIMEOUT_S
    while True:
        if time.monotonic() > deadline:
            raise TimeoutError(f"Timed out waiting for the excel write lock on '{path}'.")
        fd = os.open(lock_path, os.O_CREAT | os.O_RDWR)
        try:
            while not _try_lock(fd):
                if time.monotonic() > deadline:
                    raise TimeoutError(f"Timed out waiting for the excel write lock on '{path}'.")
                time.sleep(_LOCK_POLL_S)
            if _is_live_lockfile(fd, lock_path):
                break
        except BaseException:
            try:
                _unlock(fd)
            except OSError:
                pass
            os.close(fd)
            raise
        _unlock(fd)
        os.close(fd)
    try:
        yield
    finally:
        if os.name == "nt":
            _unlock(fd)
            os.close(fd)
            try:
                os.unlink(lock_path)
            except OSError:
                pass
        else:
            try:
                os.unlink(lock_path)
            except OSError:
                pass
            _unlock(fd)
            os.close(fd)


def resolve_excel_write_mode(write_mode: str | None) -> str:
    """Map a persisted write_mode onto a supported excel mode, defaulting to ``overwrite``.

    The UI, the frame API and hand-written YAML have all been able to persist values excel never
    honoured ("new file", "append", "error", missing); they have always behaved as overwrite and
    must keep doing so.
    """
    return write_mode if write_mode in EXCEL_WRITE_MODES else "overwrite"


def _excel_safe_frame(df: pl.DataFrame) -> pl.DataFrame:
    """Coerce the dtypes openpyxl rejects into the values ``write_excel`` produces for them.

    Nested values go through Python ``str()`` — polars refuses to cast List/Struct to Utf8, and
    ``str()`` reproduces xlsxwriter's rendering exactly. tz-aware datetimes are made naive because
    ``write_excel`` refuses them outright. Deliberately applied on the update path only: overwrite
    stays polars-verbatim.
    """
    columns: list[pl.Series | pl.Expr] = []
    for name, dtype in df.schema.items():
        if isinstance(dtype, _NESTED_DTYPES):
            values = [None if value is None else str(value) for value in df[name].to_list()]
            columns.append(pl.Series(name, values, dtype=pl.Utf8))
        elif isinstance(dtype, pl.Datetime) and dtype.time_zone is not None:
            columns.append(pl.col(name).dt.replace_time_zone(None))
    return df.with_columns(columns) if columns else df


def _next_table_name(wb: Workbook) -> str:
    """Pick a ``FrameN`` display name unused anywhere in the workbook (Excel requires global uniqueness)."""
    used = {name for worksheet in wb.worksheets for name in worksheet.tables}
    index = 0
    while f"Frame{index}" in used:
        index += 1
    return f"Frame{index}"


def _save_atomically(wb: Workbook, path: str) -> None:
    """Save *wb* over *path* via a temp file + ``os.replace`` so a killed process cannot destroy it.

    The temp name is deterministic and only ever touched under the workbook lock, so an orphan
    left by a killed writer is bounded to one per workbook and reclaimed by the next write.
    """
    tmp_path = path + ".tmp"
    try:
        wb.save(tmp_path)
        shutil.copymode(path, tmp_path)
        os.replace(tmp_path, path)  # openpyxl truncates in place; the worker's cancel path terminates()
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def _write_fresh_atomically(df: pl.DataFrame, path: str, sheet: str) -> None:
    """Write a brand-new workbook via a temp file + ``os.replace`` (same bounded temp-name scheme)."""
    tmp_path = path + ".tmp"
    try:
        df.write_excel(tmp_path, worksheet=sheet)
        os.replace(tmp_path, path)
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def _update_sheet(df: pl.DataFrame, path: str, sheet_name: str) -> None:
    """Replace *sheet_name* inside the existing workbook at *path*, preserving everything else."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = load_workbook(path)  # default data_only=False; data_only=True would freeze every formula
    if sheet_name in wb.sheetnames:
        index = wb.sheetnames.index(sheet_name)
        del wb[sheet_name]
        ws = wb.create_sheet(sheet_name, index)  # index is mandatory, else the tab moves to the end
    else:
        ws = wb.create_sheet(sheet_name)
    frame = _excel_safe_frame(df)
    ws.append(frame.columns)
    for row in frame.iter_rows():
        ws.append(row)
    if frame.height and frame.width:
        ref = f"A1:{get_column_letter(frame.width)}{frame.height + 1}"
        table = Table(displayName=_next_table_name(wb), ref=ref)
        table.tableStyleInfo = TableStyleInfo(showRowStripes=True)
        try:
            ws.add_table(table)
        except Exception:
            pass  # purely cosmetic parity with write_excel; never fail the write over it
    _save_atomically(wb, path)


def write_excel_output(
    df: pl.DataFrame,
    path: str,
    sheet_name: str | None = None,
    write_mode: str = "overwrite",
) -> None:
    """Write ``df`` to the excel file at *path*.

    ``overwrite`` replaces the whole workbook, ``create`` refuses to touch an existing file, and
    ``update`` replaces only ``sheet_name`` while keeping every other sheet — and its formatting,
    formulas and layout — intact. ``update`` against a missing file writes a fresh workbook.

    ``update`` and ``create`` hold the per-workbook lock, so parallel nodes filling different
    sheets of one file merge sequentially and a concurrent ``create`` fails instead of racing;
    both resolve symlinks first so every spelling of one workbook shares one lock and writes
    through to the real file. Mixing an ``overwrite`` node with locked writers on one file is
    unsupported — the unlocked overwrite can make a concurrent update fail loudly (BadZipFile).
    """
    sheet = sheet_name or "Sheet1"
    mode = resolve_excel_write_mode(write_mode)
    if mode == "overwrite":
        df.write_excel(path, worksheet=sheet)
        return
    path = os.path.realpath(path)
    with _workbook_lock(path):
        exists = os.path.exists(path)
        if mode == "create":
            if exists:
                raise FileExistsError(f"Cannot write '{path}': the file already exists (write mode 'create').")
            _write_fresh_atomically(df, path, sheet)
            return
        if exists:
            _update_sheet(df, path, sheet)
        else:
            _write_fresh_atomically(df, path, sheet)
